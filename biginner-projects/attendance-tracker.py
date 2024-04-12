import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

book = openpyxl.load_workbook('H://python-projects//biginner-projects//attendance.xlsx')
sheet = book['Sheet1']
r = sheet.max_row
resp = 1

c = sheet.max_column

l1 = []
l2 = ''
l3 = []

staff_mails = ['swamirahul7850@gmail.com','ankurswami90@gmail.com']

m1 = 'warning!!! you can take only one more day leave for CI class'
m2 = 'warning!!! you can take only one more day leave for Python class'
m3 = 'warning!!! you can take only one more day leave for DM class'

def savefile():
    book.save(r'H://python-projects//biginner-projects//attendance.xlsx')
    print('saved!!')

def check(no_of_days,row_num,b):
    global staff_mails
    global l2
    global l3

    for student in range(0,len(row_num)):
        if no_of_days[student] is 2:
            if b is 1:
                l1.append(sheet.cell(row=row_num[student],column=2).value)
                mailstu(l1,m1)

            elif b is 2:
                l1.append(sheet.cell(row=row_num[student],column=2).value)
                mailstu(l1,m2)

            else:
                l1.append(sheet.cell(row=row_num[student],column=2).value)
                mailstu(l1,m3)

        elif no_of_days[student] >2:
            if b is 1:
                l2 = l2+str(sheet.cell(row=row_num[student],column=1).value)

                l3.append(sheet.cell(row=row_num[student],column=2).value)
                subject = 'CI'
        
            elif b is 2 :
                l2 = l2+str(sheet.cell(row=row_num[student],column=1).value)

                l3.append(sheet.cell(row=row_num[student],column=1).value)
                subject = 'Python'

            else:
                l2 = l2+str(sheet.cell(row=row_num[student],column=1).value)

                l3.append(sheet.cell(row=row_num[student],column=1).value)
                subject = 'Data mining'
        if l2 != '' and len(l3) != 0:
            msg1 = 'you have lack of attendance in' +subject+'!!!'
            msg2 = 'the following students have lack of attendance in your subject :'+l2

            mailstu(l3,msg1)
            staff_id = staff_mails[b-1]
            mailstaff(staff_id,msg2)







def mailstu(li,msg):
    from_id = 'crazygirlaks@gmail.com'
    pwd = 'ERAkshaya485'
    s = smtplib.SMTP('smtp.gmail.com',587,timeout=120)
    s.starttls()
    s.login(from_id,pwd)

    for i in range(0,len(li)):
        to_id = li[i]
        message = MIMEMultipart()
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg,'plain'))
        content = message.as_string()
        s.sendmail(from_id,to_id,content)
        s.quit()
    print('mail sent to students')

def mailstaff(mail_id,msg):
    from_id = 'crazygirlaks@gmail.com'
    pwd = 'ERAkshaya485'
    to_id = mail_id
    message = MIMEMultipart()
    message.attach(MIMEText(msg,'plain'))
    s = smtplib.SMTP('smtp.gmail.com',578,timeout=120)
    s.starttls()
    s.login(from_id,pwd)
    content = message.as_string()
    s.sendmail(from_id,to_id, content)
    s.quit()
    print('Mail Sent to staff')

while resp is 1:
    print('1--->CI\n2--->Python\n3--->DM')
    y = int(input('Enter subject number ::'))
    no_of_absentees = int(input('number.of.absenteen ::'))
    if no_of_absentees >1:
        x = list(map(int,(input('roll no..::').split(' '))))
    else:
        x = [int(input('roll no ::'))]


    row_num = []
    
    no_of_days = []

    for student in x:

        for i in range(2,r+1):
            if y is 1:
                if sheet.cell(row=i,column=1).value is student:
                    m = sheet.cell(row=i,column=3).value
                    m = m+1
                    sheet.cell(row=i,column=3).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
            
            elif y is 2:
                if sheet.cell(row=i,column=1).value is student:
                    m = sheet.cell(row=i,column=4).value
                    m = m+1
                    sheet.cell(row=i,column=4).value = m+1
                    no_of_days.append(m)
                    row_num.append(i)

            elif y is 3:
                if sheet.cell(row=i,column=1).value is student:
                    m = sheet.cell(row=i,column=5).value
                    m = m+1
                    sheet.cell(row=i,column=5).value = m+1
                    no_of_days.append(m)
                    row_num.append(i)
    check(no_of_days,row_num,y)
    resp = int(input('another subject ? 1---->yes 0---->no'))